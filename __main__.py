from selenium import webdriver
from time import sleep
import openpyxl
from openpyxl import load_workbook
class main():

    def __init__(self,url):
        """ИНИЦИАЛИЗАЦИЯ"""
        self.driver = webdriver.Chrome()
        self.url = url
        self.urls_and_pages = {}
        self.all_titles_for_excel = ['Артикул','Наименование товара','Категория','Марка','Остаток','Продажи','Выручка','Цены','URL']

    def xlsx_append(self):
        """ДОБАВЛЕНИЕ ТАЙТЛОВ"""
        wb = load_workbook('parsing2.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')

        for i in range(1,10):
            sheet.cell(row=1,column=i).value = self.all_titles_for_excel[i-1]
        wb.save('parsing2.xlsx')

    def parse_pages(self):
        """ПОЛУЧЕНИЕ ЭЛЕМЕНТОВ СО СТРАНИЦ"""
        global k 
        wb = load_workbook('parsing2.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        self.driver.get(self.url)
        all_pages = self.driver.find_elements_by_xpath('//a[@class="paginator-item "]')
        try:
            for page in range (1,int(all_pages[-1].text)+1):
                
                self.driver.get(self.url + '&page=' + str(page))
                print(self.url + '&page=' + str(page))
                titles = self.driver.find_elements_by_class_name('plp-item__info__title')
                articul = self.driver.find_elements_by_class_name('plp-item-description__value')
                
                for title in titles:
                    print(title.text)
                    max_row = sheet.max_row
                    sheet.cell(row=max_row+1,column=2).value = title.text
                    sheet.cell(row=max_row+1,column=3).value = 'Фотообои'
                for art in articul:
                    k+=1
                    sheet.cell(row=k,column=1).value = art.text

        except:
            for page in range (1,2):
                
                self.driver.get(self.url + '&page=' + str(page))
                print(self.url + '&page=' + str(page))
                titles = self.driver.find_elements_by_class_name('plp-item__info__title')
                articul = self.driver.find_elements_by_class_name('plp-item-description__value')
                for title in titles:
                    print(title.text)
                    max_row = sheet.max_row
                    sheet.cell(row=max_row+1,column=2).value = title.text
                    sheet.cell(row=max_row+1,column=3).value = 'Фотообои'

                    for art in articul:
                        k+=1
                        sheet.cell(row=k,column=1).value = art.text

                page +=1
        wb.save('parsing2.xlsx')


def check_marks():
    global marks
    marks = []
    with open('marks','r') as f:
        for line in f:
            marks.append(line.replace('\n', ''))
check_marks()

root=main('https://leroymerlin.ru/catalogue/fotooboi/?06575={}&sortby=8')
root.xlsx_append()
k = 0
for mark in marks:
    root=main('https://leroymerlin.ru/catalogue/fotooboi/?06575={}&sortby=8'.format(mark))
    root.parse_pages()

